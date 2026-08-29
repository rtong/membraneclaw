from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
import unicodedata
from typing import Any, Iterable

from openpyxl import load_workbook

from .io_utils import read_json, sha256_file, stable_hash, utc_now, write_json


SHEET_ROLE_ALIASES = {
    "question": ("题目_Q", "Question", "Question_Q"),
    "answer": (
        "分步答案_A",
        "Gold_Answer",
        "Gold_Answer_A",
        "Stepwise Answer_A",
        "Stepwise_Gold_Answer",
    ),
    "rubric": (
        "评价标准",
        "Trajectory_Rubric",
        "Trajectory Rubric",
        "Evaluation_Rubric",
    ),
    "tool_efficiency": ("Tool效率评价", "Tool_Efficiency", "Tool Efficiency"),
}

FILENAME_REPLACEMENTS = {
    "进水量": "feed-flow",
    "进水温度": "feed-temperature",
    "进水盐度": "feed-salinity",
}

REFERENCE_TRAJECTORY_EFFICIENCY_DIMENSIONS = (
    (
        "E1",
        "E1 Tool selection and necessity",
        "Choose simulator and metadata tools that match the requested outputs.",
        "Every selected tool is necessary and scientifically appropriate.",
        "The main tools are correct, with limited unnecessary metadata or exploratory calls.",
        "Wrong simulator scope, missing required tool types, or many irrelevant calls.",
    ),
    (
        "E2",
        "E2 Information gain and call ordering",
        "Order calls so each stage or branch resolves a decision boundary.",
        "Call order follows the reference dependencies or an equally informative alternative.",
        "Mostly useful ordering with minor low-information calls.",
        "Dense or serial enumeration without regard to dependencies or information gain.",
    ),
    (
        "E3",
        "E3 State lineage and adaptive iteration",
        "Preserve candidate identity and exact upstream outputs across coupled calls.",
        "Downstream calls inherit the correct candidate state and adapt to observed results.",
        "Lineage is mostly preserved, with minor rounding or weak adaptation.",
        "Mixes candidates, substitutes target values for actual outputs, or ignores prior results.",
    ),
    (
        "E4",
        "E4 Parallelization, pruning and redundancy control",
        "Parallelize independent branches, prune resolved candidates, and avoid repetition.",
        "Independent work is parallelized and redundant/equivalent calls are avoided.",
        "Some avoidable serialization or repetition, without changing the conclusion.",
        "Unnecessary serial execution, repeated equivalent calls, or no pruning.",
    ),
    (
        "E5",
        "E5 Convergence, stopping and evidence sufficiency",
        "Stop after constraints and ranking are supported by auditable evidence.",
        "Converges within the reference budget or justifies an efficient alternative budget.",
        "Minor extra calls or a weakly justified sensitivity check.",
        "Stops too early to support the answer or continues after the decision is resolved.",
    ),
)


@dataclass(frozen=True)
class BenchmarkSource:
    case_id: str
    path: Path
    task_family: str
    benchmark_overlay: dict[str, Any] | None = None


def _display(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return format(value, ".15g")
    return str(value).strip()


def sheet_to_text(sheet) -> str:
    """Render populated worksheet rows as stable, compact model-readable text."""
    lines: list[str] = []
    last_content = ""
    for row in sheet.iter_rows(values_only=True):
        cells = [_display(value) for value in row]
        while cells and not cells[-1]:
            cells.pop()
        nonempty = [cell for cell in cells if cell]
        if not nonempty:
            if lines and lines[-1] != "":
                lines.append("")
            continue
        compact_cells: list[str] = []
        for cell in nonempty:
            if not compact_cells or cell != compact_cells[-1]:
                compact_cells.append(cell)
        line = " | ".join(compact_cells)
        if line == last_content:
            continue
        lines.append(line)
        last_content = line
    while lines and not lines[-1]:
        lines.pop()
    return "\n".join(lines)


def _first_nonempty(sheet) -> str:
    for row in sheet.iter_rows(values_only=True):
        for value in row:
            text = _display(value)
            if text:
                return text
    return sheet.title



def _normalized_header(value: Any) -> str:
    text = unicodedata.normalize("NFKC", _display(value)).casefold()
    return "".join(character for character in text if character.isalnum())


def _header_index(header_values: list[str], aliases: tuple[str, ...]) -> int | None:
    normalized = {_normalized_header(value): index for index, value in enumerate(header_values) if value}
    for alias in aliases:
        index = normalized.get(_normalized_header(alias))
        if index is not None:
            return index
    return None


def _cell(row: tuple[Any, ...], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    return _display(row[index])


def _resolve_sheet_roles(workbook, filename: str) -> dict[str, Any]:
    result: dict[str, Any] = {}
    missing: list[str] = []
    for role, aliases in SHEET_ROLE_ALIASES.items():
        name = next((candidate for candidate in aliases if candidate in workbook.sheetnames), None)
        if name is None:
            missing.append(f"{role} ({', '.join(aliases)})")
        else:
            result[role] = workbook[name]
    if missing:
        raise ValueError(f"{filename}: missing benchmark sheet roles {missing}")
    return result


def _source_stem(path: Path) -> str:
    stem = path.stem
    stem = re.sub(r"(?i)(?:[_ -](?:EN|English))$", "", stem)
    stem = re.sub(r"(?i)^D1[_-]a1(?=[_-]|$)", "D1_1a", stem)
    for source, replacement in FILENAME_REPLACEMENTS.items():
        stem = stem.replace(source, replacement)
    return stem


def derive_case_id(path: Path) -> str:
    stem = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", "-", _source_stem(path))
    slug = re.sub(r"[^A-Za-z0-9]+", "-", unicodedata.normalize("NFKC", stem)).strip("-").lower()
    slug = re.sub(r"^d(\d+)-", lambda match: f"D{match.group(1)}-", slug)
    if not slug:
        raise ValueError(f"Cannot derive case_id from benchmark filename: {path.name}")
    return slug


def derive_task_family(path: Path) -> str:
    match = re.match(r"(?i)^D(\d+)[_-](\d+[a-z])(?:[_-]|$)", _source_stem(path))
    if not match:
        raise ValueError(f"Cannot derive task_family from benchmark filename: {path.name}")
    return f"d{match.group(1)}_{match.group(2).lower()}"

def _parse_rubric(sheet) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    label_aliases = ("步骤", "能力层", "Trajectory维度", "Trajectory dimension", "Dimension")
    score_aliases = ("分值", "权重", "Weight", "Points")
    header_row = None
    header_values: list[str] = []
    label_idx = None
    score_idx = None
    for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        values = [_display(value) for value in row]
        candidate_label = _header_index(values, label_aliases)
        candidate_score = _header_index(values, score_aliases)
        if candidate_label is not None and candidate_score is not None:
            header_row = idx
            header_values = values
            label_idx = candidate_label
            score_idx = candidate_score
            break
    if header_row is None or label_idx is None or score_idx is None:
        raise ValueError(f"{sheet.title}: cannot locate rubric header row")

    focus_idx = _header_index(header_values, ("评价重点", "Evaluation focus", "Evaluation question", "Focus"))
    full_idx = _header_index(
        header_values,
        (
            "本题满分证据",
            "满分表现",
            "Full-credit evidence",
            "Full-credit evidence for this task",
            "Full credit",
        ),
    )
    partial_idx = _header_index(header_values, ("部分得分", "Partial credit"))
    failures_idx = _header_index(
        header_values,
        (
            "关键失败模式",
            "Critical failure modes",
            "Critical failure mode",
            "Key failure modes",
            "Key failure mode",
            "Zero/major error",
            "Low/zero",
        ),
    )

    steps: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        raw_id = _cell(row, label_idx)
        score = row[score_idx] if len(row) > score_idx else None
        if not raw_id or not isinstance(score, (int, float)):
            if steps:
                break
            continue
        full_evidence = _cell(row, full_idx)
        partial_credit = _cell(row, partial_idx)
        steps.append(
            {
                "step_id": len(steps) + 1,
                "step_label": raw_id,
                "reasoning_or_calculation": _cell(row, focus_idx),
                "inputs": partial_credit,
                "key_outputs": full_evidence,
                "full_credit": full_evidence,
                "common_failures": _cell(row, failures_idx),
                "max_score": float(score),
            }
        )

    if not steps:
        raise ValueError(f"{sheet.title}: rubric is empty")
    return steps, []


def _parse_tool_efficiency_rubric(sheet) -> list[dict[str, Any]]:
    label_aliases = ("效率维度", "Efficiency dimension", "Dimension")
    score_aliases = ("分值", "Points", "Weight", "Score")
    header_row = None
    header_values: list[str] = []
    label_idx = None
    score_idx = None
    for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        values = [_display(value) for value in row]
        candidate_label = _header_index(values, label_aliases)
        candidate_score = _header_index(values, score_aliases)
        if candidate_label is not None and candidate_score is not None:
            header_row = idx
            header_values = values
            label_idx = candidate_label
            score_idx = candidate_score
            break
    if header_row is None or label_idx is None or score_idx is None:
        raise ValueError(f"{sheet.title}: cannot locate tool-efficiency header row")

    focus_idx = _header_index(header_values, ("评价重点", "Evaluation focus", "Evaluation question", "Focus"))
    full_idx = _header_index(
        header_values,
        ("满分表现", "Full-credit performance", "Full-credit behavior", "Full credit"),
    )
    medium_idx = _header_index(
        header_values,
        ("中等表现", "Moderate performance", "Moderate behavior", "Medium-credit behavior", "Medium behavior", "Mid-level performance", "Partial credit"),
    )
    low_idx = _header_index(
        header_values,
        ("低分表现", "Low performance", "Low-score behavior", "Low-credit behavior", "Low-level performance", "Low/zero"),
    )
    observation_idx = _header_index(
        header_values,
        ("本题观察点", "Task-specific evidence", "Task-specific observation", "Reference evidence"),
    )

    dimensions: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        raw_label = _cell(row, label_idx)
        score = row[score_idx] if len(row) > score_idx else None
        if not raw_label or not isinstance(score, (int, float)):
            if dimensions:
                break
            continue
        dimensions.append(
            {
                "dimension_id": raw_label.split(maxsplit=1)[0],
                "dimension_label": raw_label,
                "evaluation_focus": _cell(row, focus_idx),
                "full_credit": _cell(row, full_idx),
                "medium_credit": _cell(row, medium_idx),
                "low_credit": _cell(row, low_idx),
                "case_observation": _cell(row, observation_idx),
                "max_score": float(score),
            }
        )
    if not dimensions:
        raise ValueError(f"{sheet.title}: tool-efficiency rubric is empty")
    return dimensions


def _tool_efficiency_rubric(sheet) -> dict[str, Any]:
    try:
        dimensions = _parse_tool_efficiency_rubric(sheet)
        return {
            "total_points": sum(item["max_score"] for item in dimensions),
            "dimensions": dimensions,
            "source_format": "scored_dimensions",
        }
    except ValueError as exc:
        if "cannot locate tool-efficiency header row" not in str(exc):
            raise
    reference_trajectory = sheet_to_text(sheet)
    normalized = _normalized_header(reference_trajectory)
    has_reference_table = (
        all(token in normalized for token in ("stage", "tool", "calls"))
        or all(token in normalized for token in ("step", "action", "calls"))
    )
    if "toolefficiency" not in normalized or not has_reference_table:
        raise ValueError(
            f"{sheet.title}: neither a scored efficiency rubric nor a recognized reference trajectory"
        )
    dimensions = [
        {
            "dimension_id": dimension_id,
            "dimension_label": label,
            "evaluation_focus": focus,
            "full_credit": full_credit,
            "medium_credit": medium_credit,
            "low_credit": low_credit,
            "case_observation": "Use the task-specific reference trajectory and call budget below.",
            "max_score": 20.0,
        }
        for dimension_id, label, focus, full_credit, medium_credit, low_credit
        in REFERENCE_TRAJECTORY_EFFICIENCY_DIMENSIONS
    ]
    return {
        "total_points": 100.0,
        "dimensions": dimensions,
        "source_format": "reference_trajectory",
        "reference_trajectory": reference_trajectory,
        "derivation_note": (
            "The source sheet defines a 100-point reference trajectory and call budget but no "
            "numeric dimension allocation. The project-standard five 20-point efficiency dimensions "
            "are applied while preserving the complete task-specific trajectory as Judge evidence."
        ),
    }


def _append_text(original: str, addition: str) -> str:
    return f"{original.rstrip()}\n\n{addition.strip()}" if addition.strip() else original


def _apply_benchmark_overlay(
    benchmark: dict[str, Any],
    overlay: dict[str, Any] | None,
) -> None:
    """Apply a versioned derived benchmark view without editing its source workbook."""
    if not overlay:
        return
    overlay_id = overlay.get("overlay_id")
    if not isinstance(overlay_id, str) or not overlay_id.strip():
        raise ValueError(f"{benchmark['case_id']}: benchmark_overlay.overlay_id is required")
    question_replace = overlay.get("question_replace")
    if question_replace is not None:
        if not isinstance(question_replace, str) or not question_replace.strip():
            raise ValueError(
                f"{benchmark['case_id']}: question_replace must be non-empty text"
            )
        benchmark["question_prompt"] = question_replace.strip()
    for target, source in (
        ("question_prompt", "question_append"),
        ("reference_answer", "reference_answer_append"),
    ):
        addition = overlay.get(source)
        if addition is not None:
            if not isinstance(addition, str) or not addition.strip():
                raise ValueError(f"{benchmark['case_id']}: {source} must be non-empty text")
            benchmark[target] = _append_text(benchmark[target], addition)

    step_appends = overlay.get("rubric_step_appends") or {}
    if not isinstance(step_appends, dict):
        raise ValueError(f"{benchmark['case_id']}: rubric_step_appends must be an object")
    steps_by_id = {
        str(step["step_id"]): step for step in benchmark["rubric"]["steps"]
    }
    allowed_step_fields = {
        "reasoning_or_calculation",
        "inputs",
        "key_outputs",
        "full_credit",
        "common_failures",
    }
    for step_id, additions in step_appends.items():
        step = steps_by_id.get(str(step_id))
        if step is None:
            raise ValueError(
                f"{benchmark['case_id']}: overlay references unknown rubric step {step_id}"
            )
        if not isinstance(additions, dict):
            raise ValueError(
                f"{benchmark['case_id']}: rubric step overlay {step_id} must be an object"
            )
        invalid_fields = sorted(set(additions) - allowed_step_fields)
        if invalid_fields:
            raise ValueError(
                f"{benchmark['case_id']}: unsupported rubric overlay fields {invalid_fields}"
            )
        for field, addition in additions.items():
            if not isinstance(addition, str) or not addition.strip():
                raise ValueError(
                    f"{benchmark['case_id']}: rubric overlay {step_id}.{field} must be text"
                )
            step[field] = _append_text(str(step.get(field) or ""), addition)

    view = overlay.get("benchmark_view")
    if view is not None:
        if not isinstance(view, dict):
            raise ValueError(f"{benchmark['case_id']}: benchmark_view must be an object")
        rag_need = view.get("rag_need")
        expected_route = view.get("expected_route")
        if rag_need not in {"R0", "R2"}:
            raise ValueError(f"{benchmark['case_id']}: rag_need must be R0 or R2")
        required_route = "skip_rag" if rag_need == "R0" else "use_rag"
        if expected_route != required_route:
            raise ValueError(
                f"{benchmark['case_id']}: {rag_need} requires expected_route={required_route}"
            )
        benchmark["benchmark_view"] = dict(view)

    rag_evidence = overlay.get("rag_evidence")
    if rag_evidence is not None:
        if not isinstance(rag_evidence, dict):
            raise ValueError(f"{benchmark['case_id']}: rag_evidence must be an object")
        benchmark["rag_evidence"] = dict(rag_evidence)
    benchmark["source"]["derived_overlay"] = {
        "overlay_id": overlay_id,
        "sha256": stable_hash(overlay),
    }


def import_workbook(source: BenchmarkSource) -> dict[str, Any]:
    if not source.path.exists():
        raise FileNotFoundError(f"Benchmark source not found: {source.path}")
    workbook = load_workbook(source.path, data_only=False, read_only=True)
    sheets = _resolve_sheet_roles(workbook, source.path.name)
    question = sheets["question"]
    answer = sheets["answer"]
    rubric_sheet = sheets["rubric"]
    tool_efficiency_sheet = sheets["tool_efficiency"]
    steps, global_criteria = _parse_rubric(rubric_sheet)
    efficiency_rubric = _tool_efficiency_rubric(tool_efficiency_sheet)

    result = {
        "schema_version": "1.0",
        "case_id": source.case_id,
        "task_family": source.task_family,
        "title": _first_nonempty(question),
        "question_prompt": sheet_to_text(question),
        "reference_answer": sheet_to_text(answer),
        "rubric": {
            "total_points": sum(step["max_score"] for step in steps),
            "steps": steps,
            "global_criteria": global_criteria,
        },
        "tool_efficiency_rubric": efficiency_rubric,
        "source": {
            "filename": source.path.name,
            "absolute_path": str(source.path.resolve()),
            "sha256": sha256_file(source.path),
            "sheets": [sheets[role].title for role in ("question", "answer", "rubric", "tool_efficiency")],
            "sheet_roles": {role: sheet.title for role, sheet in sheets.items()},
            "imported_at": utc_now(),
        },
    }
    _apply_benchmark_overlay(result, source.benchmark_overlay)
    validate_benchmark(result)
    return result


def validate_benchmark(benchmark: dict[str, Any]) -> None:
    required = (
        "case_id",
        "task_family",
        "title",
        "question_prompt",
        "reference_answer",
        "rubric",
        "tool_efficiency_rubric",
    )
    missing = [key for key in required if not benchmark.get(key)]
    if missing:
        raise ValueError(f"Benchmark missing required fields: {missing}")
    steps = benchmark["rubric"].get("steps") or []
    if not steps:
        raise ValueError(f"{benchmark['case_id']}: rubric has no steps")
    ids = [step["step_id"] for step in steps]
    if ids != list(range(1, len(ids) + 1)):
        raise ValueError(f"{benchmark['case_id']}: rubric steps are not consecutive: {ids}")
    total = sum(float(step["max_score"]) for step in steps)
    declared = float(benchmark["rubric"].get("total_points", 0))
    if abs(total - declared) > 1e-9:
        raise ValueError(f"{benchmark['case_id']}: rubric total mismatch {total} != {declared}")
    if abs(total - 100.0) > 1e-9:
        raise ValueError(f"{benchmark['case_id']}: expected a 100-point rubric, got {total}")


    efficiency = benchmark["tool_efficiency_rubric"]
    dimensions = efficiency.get("dimensions") or []
    if not dimensions:
        raise ValueError(f"{benchmark['case_id']}: tool-efficiency rubric is empty")
    dimension_ids = [item["dimension_id"] for item in dimensions]
    expected_ids = [f"E{index}" for index in range(1, len(dimensions) + 1)]
    if dimension_ids != expected_ids:
        raise ValueError(
            f"{benchmark['case_id']}: unexpected tool-efficiency dimensions {dimension_ids}"
        )
    efficiency_total = sum(float(item["max_score"]) for item in dimensions)
    declared_efficiency_total = float(efficiency.get("total_points", 0))
    if abs(efficiency_total - declared_efficiency_total) > 1e-9:
        raise ValueError(
            f"{benchmark['case_id']}: tool-efficiency total mismatch "
            f"{efficiency_total} != {declared_efficiency_total}"
        )
    if abs(efficiency_total - 100.0) > 1e-9:
        raise ValueError(
            f"{benchmark['case_id']}: expected a 100-point tool-efficiency rubric, "
            f"got {efficiency_total}"
        )

def load_sources(config_path: Path, project_root: Path) -> list[BenchmarkSource]:
    config = read_json(config_path)
    items = list(config.get("sources") or [])
    discovery = config.get("discovery")
    if discovery:
        root_raw = Path(discovery["root"])
        discovery_root = root_raw if root_raw.is_absolute() else project_root / root_raw
        domains = {str(value).casefold() for value in discovery.get("domains", [])}
        extensions = {str(value).casefold() for value in discovery.get("extensions", [".xlsx"])}
        ignore_prefixes = tuple(discovery.get("ignore_prefixes", ["~$"]))
        overrides = config.get("overrides") or {}
        if not discovery_root.exists():
            raise FileNotFoundError(f"Benchmark discovery root not found: {discovery_root}")
        explicit_paths = {Path(item["path"]).as_posix().casefold() for item in items}
        for source_path in sorted(discovery_root.rglob("*"), key=lambda value: value.as_posix().casefold()):
            if not source_path.is_file() or source_path.suffix.casefold() not in extensions:
                continue
            if source_path.name.startswith(ignore_prefixes):
                continue
            relative_to_discovery = source_path.relative_to(discovery_root)
            if domains and (not relative_to_discovery.parts or relative_to_discovery.parts[0].casefold() not in domains):
                continue
            relative_path = source_path.relative_to(project_root).as_posix()
            if relative_path.casefold() in explicit_paths:
                continue
            override = overrides.get(relative_path) or {}
            items.append(
                {
                    "case_id": override.get("case_id") or derive_case_id(source_path),
                    "path": relative_path,
                    "task_family": override.get("task_family") or derive_task_family(source_path),
                    "benchmark_overlay": override.get("benchmark_overlay"),
                }
            )

    sources = []
    seen_case_ids: set[str] = set()
    seen_paths: set[Path] = set()
    for item in items:
        case_id = item["case_id"]
        if case_id in seen_case_ids:
            raise ValueError(f"Duplicate benchmark case_id in {config_path}: {case_id}")
        seen_case_ids.add(case_id)
        raw_path = Path(item["path"])
        source_path = (raw_path if raw_path.is_absolute() else project_root / raw_path).resolve()
        if source_path in seen_paths:
            raise ValueError(f"Duplicate benchmark source path in {config_path}: {source_path}")
        seen_paths.add(source_path)
        overlay = item.get("benchmark_overlay")
        if overlay is not None and not isinstance(overlay, dict):
            raise ValueError(f"{case_id}: benchmark_overlay must be an object")
        sources.append(
            BenchmarkSource(
                case_id=case_id,
                path=source_path,
                task_family=item["task_family"],
                benchmark_overlay=overlay,
            )
        )
    if not sources:
        raise ValueError(f"No benchmark sources configured in {config_path}")
    return sources


def import_all(config_path: Path, output_dir: Path, project_root: Path) -> list[dict[str, Any]]:
    output_dir.mkdir(parents=True, exist_ok=True)
    previous_files: set[str] = set()
    index_path = output_dir / "index.json"
    if index_path.exists():
        previous_files = {
            str(item.get("file"))
            for item in read_json(index_path).get("cases", [])
            if item.get("file")
        }
    benchmarks = []
    for source in load_sources(config_path, project_root):
        try:
            benchmarks.append(import_workbook(source))
        except Exception as exc:
            raise ValueError(
                f"Failed to import {source.case_id} from {source.path}: {exc}"
            ) from exc
    current_files = {f"{benchmark['case_id']}.json" for benchmark in benchmarks}
    for benchmark in benchmarks:
        write_json(output_dir / f"{benchmark['case_id']}.json", benchmark)
    safe_output = output_dir.resolve()
    for stale_name in sorted(previous_files - current_files):
        stale_path = (output_dir / stale_name).resolve()
        if stale_path.parent == safe_output and stale_path.suffix == ".json" and stale_path.name != "index.json":
            stale_path.unlink(missing_ok=True)
    write_json(
        index_path,
        {
            "schema_version": "1.0",
            "generated_at": utc_now(),
            "cases": [
                {
                    "case_id": item["case_id"],
                    "task_family": item["task_family"],
                    "title": item["title"],
                    "file": f"{item['case_id']}.json",
                    "source_sha256": item["source"]["sha256"],
                }
                for item in benchmarks
            ],
        },
    )
    return benchmarks


def iter_benchmarks(directory: Path) -> Iterable[dict[str, Any]]:
    index = read_json(directory / "index.json")
    for case in index.get("cases", []):
        benchmark = read_json(directory / case["file"])
        validate_benchmark(benchmark)
        yield benchmark
