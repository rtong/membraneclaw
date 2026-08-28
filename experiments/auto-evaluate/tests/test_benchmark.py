from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from auto_evaluate.benchmark import (  # noqa: E402
    derive_case_id,
    derive_task_family,
    import_all,
    iter_benchmarks,
    load_sources,
    sheet_to_text,
)


class BenchmarkImportTests(unittest.TestCase):
    def test_sheet_serialization_compacts_repeated_merged_area_values(self):
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Title", "Title", "Title", "Title"])
        sheet.append(["Title", "Title", "Title", "Title"])
        sheet.append([None, None, None, None])
        sheet.append(["Parameter", "Parameter", "Value", "Value"])
        self.assertEqual("Title\n\nParameter | Value", sheet_to_text(sheet))

    def test_auto_discovery_covers_new_domains_and_ignores_excel_lock_files(self):
        sources = load_sources(ROOT / "configs/benchmarks.json", ROOT)
        self.assertGreaterEqual(len(sources), 117)
        self.assertTrue(all(not source.path.name.startswith("~$") for source in sources))
        families = {source.task_family for source in sources}
        self.assertTrue({"d1_1a", "d2_2a", "d3_3a", "d3_3b", "d4_4a", "d5_5a", "d6_6a", "d6_6b", "d6_6c"} <= families)
        self.assertEqual(len(sources), len({source.case_id for source in sources}))

    def test_d7_mock_reuses_six_existing_workbooks_without_copying_sources(self):
        sources = load_sources(ROOT / "configs/benchmarks_d7_mock.json", ROOT)
        self.assertEqual(6, len(sources))
        self.assertEqual(
            {f"D7-mock-0{index}" for index in range(1, 7)},
            {"-".join(source.case_id.split("-")[:3]) for source in sources},
        )
        self.assertTrue(all(source.path.exists() for source in sources))
        self.assertTrue(all(source.task_family == "d7_mock" for source in sources))
        with tempfile.TemporaryDirectory() as tmp:
            imported = import_all(
                ROOT / "configs/benchmarks_d7_mock.json",
                Path(tmp) / "normalized_d7_mock",
                ROOT,
            )
        routes = [item["benchmark_view"]["expected_route"] for item in imported]
        self.assertEqual(3, routes.count("use_rag"))
        self.assertEqual(3, routes.count("skip_rag"))
        self.assertTrue(
            all(item.get("rag_evidence") for item in imported[:3])
        )
        self.assertTrue(
            all(item["source"].get("derived_overlay") for item in imported)
        )

    def test_d7_mock_router_probe_has_three_short_matched_route_pairs(self):
        sources = load_sources(
            ROOT / "configs/benchmarks_d7_mock_router_probe.json", ROOT
        )
        self.assertEqual(6, len(sources))
        self.assertEqual(6, len({source.path for source in sources}))
        self.assertTrue(
            all(source.task_family == "d7_mock_router_probe" for source in sources)
        )
        with tempfile.TemporaryDirectory() as tmp:
            imported = import_all(
                ROOT / "configs/benchmarks_d7_mock_router_probe.json",
                Path(tmp) / "normalized_d7_mock_router_probe",
                ROOT,
            )

        routes = [item["benchmark_view"]["expected_route"] for item in imported]
        self.assertEqual(3, routes.count("use_rag"))
        self.assertEqual(3, routes.count("skip_rag"))
        self.assertTrue(all(len(item["question_prompt"]) < 700 for item in imported))
        for missing, supplied in zip(imported[0::2], imported[1::2]):
            self.assertIn("not provided", missing["question_prompt"])
            self.assertNotIn("not provided", supplied["question_prompt"])
            self.assertEqual("use_rag", missing["benchmark_view"]["expected_route"])
            self.assertEqual("skip_rag", supplied["benchmark_view"]["expected_route"])
            self.assertTrue(missing.get("rag_evidence"))

    def test_d7_mock_router_natural_removes_explicit_missing_knowledge_cues(self):
        sources = load_sources(
            ROOT / "configs/benchmarks_d7_mock_router_natural.json", ROOT
        )
        self.assertEqual(6, len(sources))
        self.assertEqual(6, len({source.path for source in sources}))
        with tempfile.TemporaryDirectory() as tmp:
            imported = import_all(
                ROOT / "configs/benchmarks_d7_mock_router_natural.json",
                Path(tmp) / "normalized_d7_mock_router_natural",
                ROOT,
            )

        routes = [item["benchmark_view"]["expected_route"] for item in imported]
        self.assertEqual(3, routes.count("use_rag"))
        self.assertEqual(3, routes.count("skip_rag"))
        forbidden_cues = (
            "not provided",
            "cannot be calculated",
            "not outputs of the process simulator",
        )
        for missing, supplied in zip(imported[0::2], imported[1::2]):
            missing_prompt = missing["question_prompt"].lower()
            self.assertTrue(all(cue not in missing_prompt for cue in forbidden_cues))
            self.assertLess(len(missing["question_prompt"]), 300)
            self.assertEqual("use_rag", missing["benchmark_view"]["expected_route"])
            self.assertEqual("skip_rag", supplied["benchmark_view"]["expected_route"])
            self.assertTrue(missing.get("rag_evidence"))

    def test_filename_metadata_derivation_is_stable(self):
        self.assertEqual("D1-1b-high-salinity-robust", derive_case_id(Path("D1_1b_HighSalinityRobust_EN.xlsx")))
        self.assertEqual("D2-2a-feed-salinity-salt-shock", derive_case_id(Path("D2_2a_进水盐度_SaltShock_EN.xlsx")))
        self.assertEqual("d1_1a", derive_task_family(Path("D1_a1_Feasibility_EN.xlsx")))
        self.assertEqual("d3_3b", derive_task_family(Path("D3_3b_N01_px_efficiency_inference_English.xlsx")))
        self.assertEqual("D6-6c-h01", derive_case_id(Path("D6_6c_H01_EN.xlsx")))
        self.assertEqual("d6_6c", derive_task_family(Path("D6_6c_H01_EN.xlsx")))

    def test_imports_all_configured_workbooks_with_100_point_rubrics(self):
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "normalized"
            imported = import_all(ROOT / "configs/benchmarks.json", output, ROOT)
            configured_sources = load_sources(ROOT / "configs/benchmarks.json", ROOT)
            self.assertEqual(len(configured_sources), len(imported))
            self.assertTrue(all(item["rubric"]["steps"] for item in imported))
            self.assertTrue(all(item["rubric"]["total_points"] == 100 for item in imported))
            self.assertTrue(all(item["tool_efficiency_rubric"]["dimensions"] for item in imported))
            self.assertTrue(
                all(item["tool_efficiency_rubric"]["total_points"] == 100 for item in imported)
            )
            self.assertTrue(
                all(
                    [row["dimension_id"] for row in item["tool_efficiency_rubric"]["dimensions"]]
                    == ["E1", "E2", "E3", "E4", "E5"]
                    for item in imported
                )
            )
            d6 = next(item for item in imported if item["case_id"] == "D6-6a-n01")
            self.assertEqual("reference_trajectory", d6["tool_efficiency_rubric"]["source_format"])
            self.assertIn("simulate_ro", d6["tool_efficiency_rubric"]["reference_trajectory"])
            loaded = list(iter_benchmarks(output))
            self.assertEqual([item["case_id"] for item in imported], [item["case_id"] for item in loaded])

    def test_question_prompt_is_serialized_only_from_the_question_sheet(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "normalized"
            imported = import_all(ROOT / "configs/benchmarks.json", output, ROOT)
            for item in imported:
                source = item["source"]
                roles = source["sheet_roles"]
                workbook = load_workbook(source["absolute_path"], data_only=False, read_only=True)
                expected_question = sheet_to_text(workbook[roles["question"]])
                self.assertEqual(expected_question, item["question_prompt"])
                self.assertNotEqual(roles["question"], roles["answer"])
                self.assertNotEqual(item["question_prompt"], sheet_to_text(workbook[roles["answer"]]))


    def test_skill_does_not_contain_pilot_reference_answers(self):
        skill_root = ROOT / "skills/swro-watertap"
        text = "\n".join(
            path.read_text(encoding="utf-8")
            for path in skill_root.rglob("*")
            if path.is_file() and path.suffix in {".md", ".json"}
        )
        for leaked_value in (
            "5.05",
            "5.10 MPa",
            "3.726",
            "396.895",
            "98.72%",
            "0.00020 mol/s",
            "0.00024 mol/s",
            "0.071423",
            "-0.051875",
        ):
            self.assertNotIn(leaked_value, text)

    def test_router_skill_v012_is_generic_and_active(self):
        skill_text = (
            ROOT / "skills/swro-rag-router/v0.1.2/SKILL.md"
        ).read_text(encoding="utf-8")
        router_config = json.loads(
            (ROOT / "configs/router_evaluation.json").read_text(encoding="utf-8")
        )
        skill_variant = next(
            row for row in router_config["variants"] if row["id"] == "router-skill"
        )
        self.assertEqual(
            "swro-rag-router@0.1.2", skill_variant["skill_version"]
        )
        self.assertIn("Apply this sufficiency gate", skill_text)
        self.assertIn("Use `FULLY_SPECIFIED_NUMERIC_TASK` only after", skill_text)
        for leaked_term in (
            "D7-natural",
            "78 bar",
            "12%",
            "48%",
            "Safety file.xlsx",
            "RO-operational Manual.pdf",
        ):
            self.assertNotIn(leaked_term, skill_text)


if __name__ == "__main__":
    unittest.main()
